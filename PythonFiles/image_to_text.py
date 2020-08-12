"""
Author: Xiaoye, Yudi
Date created: 7/13/2020
Date last modified: 8/06/2020
Python Version: 3.8
Description: this script is used to extract text from table cell images and format them
into excel/json format
"""

import os
import re
from functools import cmp_to_key
import cv2
import pytesseract
import numpy as np
import openpyxl
import concurrent.futures
import Levenshtein
import json
from objdict import ObjDict
import warnings

warnings.filterwarnings("ignore")

# If you are windows user, please comment this line:
T_PATH = None

# And then uncomment and change the following T_PATH, the T_PATH is the tesseract executable in your PATH
# T_PATH = r'D:\Program Files\Tesseract-OCR\tesseract'

os.environ['OMP_THREAD_LIMIT'] = '1'

# the INPUT_NPY_PATH path is where you store the output of cell's coordinates
INPUT_NPY_PATH = r'/home/ubuntu/webserver/coordinates'

# stores current file name
FILE_TO_READ = ''

# the OUTPUT_EXCEL_PATH path is where you want to store the excel output table
OUTPUT_EXCEL_PATH = r'/home/ubuntu/webserver/excel'

# the OUTPUT_EXCEL_PATH path is where you want to store the json output
OUTPUT_JSON_PATH = r'/home/ubuntu/webserver/json'

# the IMAGE_PATH is where you store the images of each tables
IMAGE_PATH = r'/home/ubuntu/webserver/outputImages/{}'


def write_to_excel(to_be_compared):
    """
    This function is used to create excel output of the table that is being processed.
    :param to_be_compared: the content of the table, a list of lines, each line contains several cells
    """
    global FILE_TO_READ

    # decide the file path
    check_file = FILE_TO_READ.split(".")[0][-1]
    if int(check_file) > 1:
        path = OUTPUT_EXCEL_PATH + '/' + FILE_TO_READ.split(".")[0][0:-1] + '1' + '.xlsx'
    else:
        path = OUTPUT_EXCEL_PATH + '/' + FILE_TO_READ.split(".")[0] + '.xlsx'

    # workbook is created or continuous write
    new_row = 0
    if os.path.isfile(path) and int(check_file) > 1:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.worksheets[0]  # generate the first sheet within the spreadsheet file
        rows_old = sheet.max_row  # get the count of existing rows of the sheet
        new_row = rows_old
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

    # write into excel by column and row
    for i in range(0, len(to_be_compared)):
        for j in range(0, len(to_be_compared[i])):
            sheet.cell(row=i + 1 + new_row, column=j + 1, value=str(to_be_compared[i][j][6]))
    workbook.save(path)


def write_to_json(to_be_compared):
    """
    This function is used to create json output of the table that is being processed.
    :param to_be_compared: the content of the table, a list of lines, each line contains several cells
    """

    # decide the file path and write into json_dict
    check_file = FILE_TO_READ.split(".")[0][-1]
    if int(check_file) > 1:
        # read previous json file and continuous write
        path = OUTPUT_JSON_PATH + '/' + FILE_TO_READ.split(".")[0][0:-1] + '1' + '.json'
        with open(path) as f:
            json_dict = json.load(f)
        prev_row_list = [int(x) for x in list(json_dict.keys())]
        start_row_num = max(prev_row_list) + 1
        json_dict = parse_json(json_dict, start_row_num, to_be_compared)
    else:
        # set the PATH and create a new ObjDict
        path = OUTPUT_JSON_PATH + '/' + FILE_TO_READ.split(".")[0] + '.json'
        json_dict = ObjDict()
        json_dict = parse_json(json_dict, 0, to_be_compared)

    # write into json
    json_object = json.dumps(json_dict, indent=4)
    with open(path, "w") as outfile:
        outfile.write(json_object)


def parse_json(json_dict, start_row_num, to_be_compared):
    """
    This function is a helper function to create json output of the table that is being processed.
    :param json_dict: a json dictionary, either be None or contains previous pages' data
    :param start_row_num: a integer representing the start row num for new input cells, either 0 or a certain number greater than 0 (if continous write)
    :param to_be_compared: the content of the table, a list of lines, each line contains several cells
    """

    # write into json dict by row number and column number
    for i in range(len(to_be_compared)):
        json_row_list = []
        for j in range(len(to_be_compared[i])):
            cell = to_be_compared[i][j]
            json_column_dict = {}
            json_cell_dict = {'type': str(cell[0]), 'x_left': str(cell[1]), 'y_up': str(cell[2]),
                              'x_right': str(cell[3]), 'y_down': str(cell[4]), 'score': str(cell[5]),
                              'text': re.sub(u"(\u2018|\u2019)", "'", cell[6])}
            json_column_dict[str(j)] = json_cell_dict
            json_row_list.append(json_column_dict)
        json_dict[str(i + start_row_num)] = json_row_list
    return json_dict


def check_column(cell, j):
    """
    This function checks which column the current cell belongs to
    : param cell: cell to be check
    : param j: current column index
    :return: adjusted index for the cell
    """

    if cell[0] == -1:
        return j
    else:
        for k in range(j + 1, sample_line.__len__()):
            if cell[3] < sample_line[k][1] + 5:
                return j
            else:
                j += 1
        return j


def find_missing_cell(table_content, col_num, img_cv):
    """
    This function matches each cell into their column, and calculate the coordinates of missing cells
    :param img_cv: the current table that is being processed
    :param table_content: A list of lines, each line contains a list of cells
    :param col_num: the column num for current table that is being processed
    """
    global sample_line

    sample_line = None
    num_of_rows = len(table_content)
    for i in range(0, num_of_rows):
        if len(table_content[i]) == col_num:
            sample_line = table_content[i]
            break

    if sample_line is None:
        return

    to_be_compared = []
    for i in range(0, num_of_rows):

        if len(table_content[i]) >= col_num:
            if table_content[i][0][0] == -1:
                row_top = table_content[i][1][2]
                row_bottom = table_content[i][1][4]
                left = 5
                if sample_line.__len__() >= 2:
                    right = sample_line[1][1] - 5
                else:
                    right = sample_line[0][3]
                table_content[i][0][6] = cell_to_text(-1, int(row_top), int(row_bottom), int(left), int(right), img_cv)
            to_be_compared.append(table_content[i])
        else:
            l = ['----'] * sample_line.__len__()
            row_top = 3000
            row_bottom = 0
            for j in range(0, len(table_content[i])):
                if j == 0 and table_content[i][j][0] == -1:
                    row_top = table_content[i][1][2]
                    row_bottom = table_content[i][1][4]
                    left = 5
                    if sample_line.__len__() >= 2:
                        right = sample_line[1][1] - 5
                    else:
                        right = sample_line[0][3]
                    table_content[i][j][6] = cell_to_text(-1, int(row_top), int(row_bottom), int(left), int(right),
                                                          img_cv)
                    l[j] = table_content[i][j]
                    continue

                row_top = min(table_content[i][j][2], row_top)
                row_bottom = max(table_content[i][j][4], row_bottom)
                correct_pos = check_column(table_content[i][j], j)
                l[correct_pos] = table_content[i][j]
                if correct_pos == j:
                    continue
                else:
                    for k in range(correct_pos - 1, j - 1, -1):
                        right = l[k + 1][1] - 5
                        left = sample_line[k][1]
                        l[k] = [-2, left, row_top, right, row_bottom, 0,
                                cell_to_text(-2, int(row_top), int(row_bottom), int(left), int(right), img_cv)]
            for j in range(correct_pos + 1, len(l)):
                left = l[j - 1][3] + 5
                right = sample_line[j][3]
                l[j] = [-2, left, row_top, right, row_bottom, 0,
                        cell_to_text(-2, int(row_top), int(row_bottom), int(left), int(right), img_cv)]
            to_be_compared.append(l)

    write_to_excel(to_be_compared)
    write_to_json(to_be_compared)


def cell_to_text(cell_type, y_up, y_down, x_left, x_right, img_cv):
    """
    This function is used to crop the cell from the whole table image using OpenCV and
    extract character within each cell using pytesseract (an OCR tool).
    :param cell_type: type of cell
    :param y_up: the upper bound of the cell
    :param y_down: the lower bound of the cell
    :param x_left: the left bound of the cell
    :param x_right: the right bound of the cell
    :param img_cv: the img of the table that is being processed
    :return: the text within the cell
    """

    if T_PATH is not None:
        pytesseract.pytesseract.tesseract_cmd = T_PATH

    # check if the coordinates of the cell are valid
    if y_up >= y_down or x_left >= x_right or y_up > img_cv.shape[0] or x_right > img_cv.shape[
        1] or y_up < 0 or y_down < 0 \
            or x_left < 0 or x_right < 0:
        return ""

    # crop the image
    img_crop = img_cv[y_up:y_down, x_left:x_right]

    # if there is no image after cropping, then return ""
    if img_crop is None:
        return ""

    # binarize img_crop
    gray = cv2.cvtColor(img_crop, cv2.COLOR_RGB2GRAY)
    ret, binary = cv2.threshold(gray, 127, 255, cv2.THRESH_BINARY)
    narr = np.array(binary)
    if len(narr.astype(np.int8)[narr == 0]) == 0:
        return ""

    # by default OpenCV stores images in BGR format and since pytesseract assumes RGB format,
    # we need to convert from BGR to RGB format/mode:
    img_rgb = cv2.cvtColor(img_crop, cv2.COLOR_BGR2RGB)

    # resize img_rgb
    scale_percent = 200  # percent of original size
    width = int(img_rgb.shape[1] * scale_percent / 100)
    height = int(img_rgb.shape[0] * scale_percent / 100)
    dim = (width, height)
    img_rgb = cv2.resize(img_rgb, dim, interpolation=cv2.INTER_CUBIC)

    # apply dilation and erosion to remove some noise
    kernel = np.ones((1, 1), np.uint8)
    img_rgb = cv2.dilate(img_rgb, kernel, iterations=1)
    img_rgb = cv2.erode(img_rgb, kernel, iterations=1)

    # read text
    if cell_type == 3:
        text = pytesseract.image_to_string(img_rgb, config='-c tessedit_char_whitelist=0123456789.,$() --psm 6')
    else:
        text = pytesseract.image_to_string(img_rgb, lang='eng', config='--psm 7 --oem 1')

    # if no text is detected, return ""
    if text.__len__() == 0:
        return ""

    # delete irregular text for specific cell (type =-2 or type =3)
    digit_cell = r'[0-9().,\$\s]*$'
    if cell_type == -2 or cell_type == 3:
        if re.match(digit_cell, text.lower().strip()) == None:
            return ""

    # adjust text of digit cell
    if cell_type == 3 or (cell_type == -2 and re.search('[a-z]', text) is not True):
        # solve "09" --> "0.9"
        index = 0
        while index < text.__len__() and text[index].isdigit() is False:
            index += 1
        if index < text.__len__() and text[index] == "0":
            if index + 1 < len(text) and text[index + 1].isdigit() is True:
                text = text[0:index + 1] + "." + text[index + 1:]

        # solve "XXX$" case
        if len(text.strip()) > 1 and text.strip()[-1] == '$':
            text = text.strip()[:-1]

        # solve "., $ 1,345,270" case
        ind = 0
        while text.__len__() > 1 and text[ind] != '$' and text[ind] != '(' and text[ind].isdigit() is False:
            text = text[1:]

        # adjust comma and period
        text_list = text.split('.')
        new_text = text_list[0]
        if len(text_list) > 1:
            for text in text_list[1:]:
                if len(text) > 2:
                    new_text += ',' + text
                else:
                    new_text += '.' + text
        text = new_text

        # adjust duplicate $
        if text.count('$') > 1:
            text = text[text.count('$') - 1:]

        text_list = list(text)
        if text.count('.') == 0:
            count = -1
            for i in range(len(text_list) - 1, -1, -1):
                if text_list[i].isdigit() is False or text_list[i] == '$':
                    break
                count += 1
                if count == 3:
                    if text_list[i] != ',':
                        text_list.insert(i + 1, ',')
                        count = 0
                    else:
                        count = -1

        # maintain () structure
        count = text.count(')') + text.count('(')
        if count == 1:
            if text_list[-1] == ')':
                if text[0] != '$':
                    text_list.insert(0, '(')
                else:
                    text_list.insert(1, '(')
            else:
                text_list.append(')')
        s = ''
        text = s.join(text_list)

    return text


def cmp_within_one_line(cell_1, cell_2):
    """
    This function is used to sort cells within a single line.
    Sorting rules: left to right
    cell data format: [cell_type,x_left, y_up, x_right, y_down]
    :param cell_1: first cell data
    :param cell_2: second cell data
    """
    if cell_1[3] < cell_2[3]:  # compare x_right
        return -1
    else:
        return 1


def clear_dup_and_solve_single_line(lines_of_table, img_cv):
    """
    This function will further clear duplicate digit cells and solve single line problem
    :param lines_of_table: A list of lists, the line structure of the table, each line contains several cells
    """
    table_content = []
    cells_num_in_single_line = []
    prev_text_list = []
    for element in lines_of_table:
        # sort data within each line
        new_order = sorted(lines_of_table[element], key=cmp_to_key(cmp_within_one_line))

        text_list = []
        prev_cell = []
        for cell in new_order:
            # if the cell type is -1, then change its text to -1
            if cell[0] == -1:
                cell[6] = -1
                text_list.append(cell)
                continue

            # delete the cell if it's only a '$'
            if cell[6].strip() == '$':
                continue

            # delete duplicate digit cell based on their intersection area
            if cell[0] == 3 and len(prev_cell) != 0 and prev_cell[0] == 3:
                if compare_dup_area(cell, prev_cell):
                    # only overwrite if its height is smaller than previous
                    if (cell[4] - cell[2]) < (prev_cell[4] - prev_cell[2]):
                        text_list.pop(-1)  # delete prev cell
                    else:
                        continue
            text_list.append(cell)

            # if it's second line then concatenate the row title with prev
            # title and drop previous text_list
            text = cell[6]
            months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
                      'November', 'December']
            # normally, a row title start with digits or lower case characters
            # or month will be a second line. we can add more into the condition
            # below after we meed more corner cases
            if len(text) > 0 and len(prev_text_list) > 0 and cell[0] == 1 and (
                    text[0].islower() or text[0].isdigit() or text.split(' ')[0] in months or re.match(r'^[0-9].*$',
                                                                                                       text.strip()) is not None):
                text_list[0][6] = str(prev_text_list[0][6]).strip() + ' ' + str(text).strip()
                table_content.pop(-1)
            prev_cell = cell

        tmp_list = [x for x in text_list if x[6]]
        if len(tmp_list) == len(text_list):
            cells_num_in_single_line.append(len(text_list))
        table_content.append(text_list)
        prev_text_list = text_list  # record the previous line in case it has second line

    # find the mode column number
    mode_list = [x for x in cells_num_in_single_line if 2 < x]
    if mode_list is None or len(mode_list) == 0:
        return
    if mode_list.count(max(mode_list)) / len(mode_list) >= 0.1:
        mode_num = max(mode_list)
    else:
        mode_num = max(set(mode_list), key=mode_list.count)
    find_missing_cell(table_content, mode_num, img_cv)


def cmp_within_whole_table(cell_1, cell_2):
    """
    This function is used to sort cells within a whole table.
    Sorting rules: top to bottom, left to right
    cell data format: [cell_type,x_left, y_up, x_right, y_down]
    :param cell_1: first cell data
    :param cell_2: second cell data
    """

    if cell_1[2] < cell_2[2]:  # first compare y_up
        return -1
    elif cell_1[2] > cell_2[2]:
        return 1
    else:  # then compare x_right
        if cell_1[3] < cell_2[3]:
            return -1
        else:
            return 1


def sort_into_lines(result):
    """
    This function is used to group sorted cells into each line.
    :param result: sorted cell data
    :return: lines of cells
    """
    # the sorted data will be grouped into each line
    lines_of_table = {}
    wait_list = []
    column_wait_list = []
    current_bottom = 0
    for cell in result:
        if cell[0] == 1:  # if this is a row title
            cells_in_line = [cell]
            current_bottom = cell[4]
            current_top = cell[2]
            no_row_title = [[-1, -1, -1, -1, -1, -1, "-1"]]
            no_row_bottom = 0
            for c in wait_list:  # handling wait_list
                if c[4] - current_top < 5:
                    if c[0] == 3:
                        no_row_bottom = no_row_bottom + c[4]
                        no_row_title.append(c)
                    else:
                        column_wait_list.append(c)
                else:
                    cells_in_line.append(c)
            if len(column_wait_list) > 0:
                top = column_wait_list[0][2]
                column_title = [column_wait_list[0]]
                lines_of_table[top] = column_title
                for col in column_wait_list[1:]:
                    if abs(top - col[2]) < 0.6 * (col[4] - col[2]):
                        lines_of_table[top].append(col)
                    else:
                        top = col[2]
                        column_title = [col]
                        lines_of_table[top] = column_title
            if no_row_title.__len__() > 1:
                lines_of_table[no_row_bottom / no_row_title.__len__()] = no_row_title
            lines_of_table[current_bottom] = cells_in_line
            wait_list = []
        else:  # have to decide which row it belongs to
            if current_bottom == 0:  # if no row has been detected, then go to wait list
                wait_list.append(cell)
            else:  # if there is one active row, check whether belongs to it or not
                if abs(current_bottom - cell[4]) < 0.6 * (cell[4] - cell[2]):
                    lines_of_table[current_bottom].append(cell)
                else:
                    wait_list.append(cell)
    return lines_of_table


def delete_dup_cells(cell_data):
    """
    Delete duplicate row title cells based on area intersection.
    :param cell_data: A list of cells, [cell1, cell2, ....]
    """

    new_cell_data = []
    for i in range(len(cell_data)):
        cell = cell_data[i]

        # delete the cell if it's only a line
        if cell[1] == cell[3] or cell[2] == cell[4]:
            continue

        # delete dup cells for type 1 and type 2 cell
        if cell[0] == 1 or cell[0] == 2:
            if len(new_cell_data) > 1:
                findDup = False
                i = -1
                while i >= max(-4, -len(
                        new_cell_data)):  # maximum look back 4 cells or not exceed the length of new_cell_data
                    prev_cell = new_cell_data[i]

                    # if completely same coordinates, delete both
                    if cell[1:5] == prev_cell[1:5]:
                        new_cell_data.pop(-1)

                    # if find duplicate cells (their interaction area
                    # is greater than half of the small cell)
                    if compare_dup_area(cell, prev_cell) and Levenshtein.ratio(cell[6], prev_cell[6]) > 0.35:
                        findDup = True
                        # only overwrite if its height is smaller than previous
                        if (cell[4] - cell[2]) < (prev_cell[4] - prev_cell[2]):
                            new_cell_data[i] = cell
                        break
                    i -= 1
                if findDup:
                    continue
        new_cell_data.append(cell)
    return new_cell_data


def compare_dup_area(cell1, cell2):
    """
    This function is used to decide if two cells are duplicate through calculating
    their intersection area with their coordinates. If the intersection area is greater
    that 0.55 * min(area_1, area_2), then we will regard the two cells are duplicate,
    and return True, else return false
    """
    A, B, C, D = cell1[1], cell1[4], cell1[3], cell1[2]
    E, F, G, H = cell2[1], cell2[4], cell2[3], cell2[2]

    # calculate the left, right, top, bottom boundary of the two cells
    left_boundary = max(A, E)
    right_boundary = min(C, G)
    top_boundary = max(D, H)
    bottom_bondary = min(B, F)

    area_1 = (abs(C - A)) * (abs(D - B))
    area_2 = (abs(G - E)) * (abs(H - F))

    if (left_boundary < right_boundary) and (bottom_bondary > top_boundary):
        # area_1 and area_2 has overlapped area
        intersection = (right_boundary - left_boundary) * (bottom_bondary - top_boundary)
    else:
        # area_1 and area_2 has no overlapped area
        intersection = 0
    return intersection > 0.55 * min(area_1, area_2)


def convert_one_cell(data):
    """
    This function is a helper function to call cell_to_text(),
    will help with the multi processing process.
    :param data: a tuple of the cell and it's image file -> (cell, img_cv)
    """

    cell, img_cv = data
    return cell_to_text(int(cell[0]), int(cell[2]), int(cell[4]), int(cell[1]), int(cell[3]), img_cv)


def convert_all_cells(cells, img_cv):
    """
    This function is used to detect text content within each cell using pytesseract,
    and multi-processing is implemented here to improve the efficiency.
    :param cells: A list of cells waiting to be detected.
    :param img_cv: The current image version file that is being processed.
    :return: A list of cells, each cell's format: [cell_type, x_left, y_up, x_right, y_bottom, score, TEXT]
    """
    with concurrent.futures.ProcessPoolExecutor(max_workers=4) as executor:
        data_list = [(cell, img_cv) for cell in cells]
        for cell, string in zip(cells, executor.map(convert_one_cell, data_list)):
            cell.append(string)

def sort_data(to_be_sort, img_cv):
    """
    This function is used to sort each table's data within a data list
    :param to_be_sort: The structure of data list is:[[cell_1], [cell_2], [cell_3] ...]
    :param img_cv: The current image version file that is being processed
    """

    result = sorted(to_be_sort, key=cmp_to_key(cmp_within_whole_table))

    # use multi-threading to detect text content within each cell
    convert_all_cells(result, img_cv)

    # delete duplicate row titles
    result = delete_dup_cells(result)

    # sort cells into according lines
    lines_table = sort_into_lines(result)

    # delete duplicate number cells and solve single line problems
    clear_dup_and_solve_single_line(lines_table, img_cv)


def load_npy():
    """
    This function is used to read cell data from .npy
    The structure of the cell_data is :[cell_1, cell_2, cell_3,...]
    The structure of each cell_i is: [cell_type,x_left, y_up, x_right, y_down,score]
    """
    cell_data = []
    arr = np.load(INPUT_NPY_PATH + '/' + FILE_TO_READ)
    label_arr = np.load(INPUT_NPY_PATH + '/' + FILE_TO_READ.split('.')[0] + '_labels.npy')

    IMAGE_ID = FILE_TO_READ.split('.')[0] + '.jpg'

    # read table image; the path is where you store the images for each table
    img_cv = cv2.imread(IMAGE_PATH.format(IMAGE_ID))

    # add image name, should be deleted after we have image id as input
    row_num = 0
    for row in arr:
        if label_arr[row_num] == 0 or row[0] == row[2] or row[1] == row[3]:
            row_num += 1
            continue
        row = row.tolist()
        row.insert(0, label_arr[row_num])  # insert cell data type
        cell_data.append(row)
        row_num += 1

    sort_data(cell_data, img_cv)


def image_to_text():
    """
    This function is used to process npy files listed in INPUT_NPY_PATH
    """
    files = os.listdir(INPUT_NPY_PATH)
    for filename in sorted(files):
        if filename != '.DS_Store':
            global FILE_TO_READ
            FILE_TO_READ = filename
            if "label" in filename:
                continue
            try:
                print("Start to transfer for {} ...".format(FILE_TO_READ))
                load_npy()
                print("Finished for {} !".format(FILE_TO_READ))
                print(" ")
            except Exception as e:
                print(e)
                print("Found error! Please check for {}".format(FILE_TO_READ))
                print(" ")


if __name__ == '__main__':
    image_to_text()
